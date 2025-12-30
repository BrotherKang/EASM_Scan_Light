import nmap
import requests
import re
import os
import glob
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

# =================設定區=================
OUTPUT_DIR = 'reports'         
MAX_WORKERS = 10                
TIMEOUT_GEO = 3                 
# =======================================

class LightEASMScannerV3:
    def __init__(self):
        self.nm = nmap.PortScanner()
        self.start_time = datetime.now()
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)

    def get_geo_info(self, ip):
        try:
            url = f"http://ip-api.com/json/{ip}?fields=status,country,city,isp"
            resp = requests.get(url, timeout=TIMEOUT_GEO).json()
            if resp.get('status') == 'success':
                return f"{resp.get('country')}-{resp.get('city')}", resp.get('isp')
        except: pass
        return "未知", "未知"

    def parse_scan_results(self, ip, scan_data):
        result = {
            "IP": ip,
            "地理位置": "未知", "ISP": "未知",
            "資產狀態": "◎ 持續監控", 
            "異動摘要": "無顯著異動",
            "開放Port": [],
            "SSL/TLS風險": [],
            "HSTS狀態": "不適用",
            "CVE漏洞": set(),     # 集合自動去重
            "vuln_details": [],  # 詳細條列
            "建議": set(),
            "raw_ports": []
        }

        geo, isp = self.get_geo_info(ip)
        result["地理位置"] = geo
        result["ISP"] = isp

        if ip not in scan_data.all_hosts():
            result["建議"].add("主機無回應")
            return result

        tcp_ports = scan_data[ip].get('tcp', {})
        if not tcp_ports:
            result["建議"].add("無開放 Port")
            return result

        web_ports_found = False
        
        for port, info in tcp_ports.items():
            if info.get('state') != 'open': continue
            
            service = info.get('name', 'unknown')
            version = info.get('version', '')
            product = info.get('product', '')
            service_full = f"{service} {product} {version}".strip()
            
            result["開放Port"].append(f"{port}/{service}")
            result["raw_ports"].append(port)

            if port in [80, 443, 8080, 8443] or 'http' in service:
                web_ports_found = True

            scripts = info.get('script', {})

            # === CVE 漏洞解析 (加入去重邏輯) ===
            if 'vulners' in scripts:
                # 1. 抓取原始字串
                raw_cves = re.findall(r'(CVE-\d{4}-\d+)', scripts['vulners'])
                
                # 2. 透過列表推導式進行「清理」與「初步去重」
                # strip() 確保沒有空白，並確保格式統一
                clean_cves = {cve.strip() for cve in raw_cves} 
                
                if clean_cves:
                    # 更新至彙整集合 (這裡 set 會再次保證 IP 層級的去重)
                    result["CVE漏洞"].update(clean_cves)
                    result["建議"].add(f"Port {port} 發現已知漏洞")
                    
                    # 更新到詳細清單 (給工程師看)
                    for cve in sorted(list(clean_cves)):
                        # 增加防重複檢查：確保同一個 Port 下不會重複列出同一個 CVE
                        if not any(d['port'] == port and d['cve'] == cve for d in result["vuln_details"]):
                            result["vuln_details"].append({
                                "port": port,
                                "service": service_full,
                                "cve": cve,
                                "desc": f"於 Port {port} 偵測到 {cve}"
                            })

            # === SSL/TLS & HSTS ===
            if 'ssl-enum-ciphers' in scripts:
                ssl_out = scripts['ssl-enum-ciphers']
                if any(x in ssl_out for x in ["SSLv2", "SSLv3"]):
                    result["SSL/TLS風險"].append(f"Port {port}: SSLv2/v3")
                if "TLSv1.0" in ssl_out:
                    result["SSL/TLS風險"].append(f"Port {port}: TLS 1.0")

            if port in [443, 8443] and 'http-security-headers' in scripts:
                if 'Strict-Transport-Security' in scripts['http-security-headers']:
                    result["HSTS狀態"] = "✅ 已啟用"
                else:
                    result["HSTS狀態"] = "❌ 未啟用"

        if web_ports_found and result["HSTS狀態"] == "不適用" and any(p in tcp_ports for p in [443, 8443]):
             result["HSTS狀態"] = "❓ 未偵測到"

        if not result["CVE漏洞"] and not result["SSL/TLS風險"]:
            result["建議"].add("無重大已知風險")

        return result

    def scan_single_ip(self, ip):
        print(f"[*] 正在掃描: {ip}")
        try:
            # 提高版本偵測強度以獲取更精準的漏洞比對
            args = "-sV --version-intensity 5 -T4 --open --script vulners,ssl-enum-ciphers,http-security-headers"
            nm_instance = nmap.PortScanner() # 每個 Thread 使用獨立實體避免衝突
            nm_instance.scan(ip, arguments=args)
            return self.parse_scan_results(ip, nm_instance)
        except Exception as e:
            return {"IP": ip, "建議": {f"掃描出錯: {str(e)}"}, "CVE漏洞": set(), "vuln_details": [], "開放Port": [], "raw_ports": [], "資產狀態": "錯誤", "異動摘要": "掃描失敗", "地理位置": "未知", "ISP": "未知", "HSTS狀態": "未知", "SSL/TLS風險": []}

    def load_history_data(self):
        files = sorted(glob.glob(os.path.join(OUTPUT_DIR, "*.xlsx")))
        files = [f for f in files if "Diff" not in f and not os.path.basename(f).startswith('~$')]
        
        if not files: return None
        
        last_file = files[-1]
        print(f"[INFO] 讀取歷史檔案: {os.path.basename(last_file)}")
        
        try:
            wb = openpyxl.load_workbook(last_file, read_only=True)
            if "詳細結果" in wb.sheetnames:
                ws = wb["詳細結果"]
            else:
                ws = wb.active 
                
            headers = [cell.value for cell in ws[1]]
            try:
                ip_idx = headers.index("IP")
                port_idx = headers.index("開放Port")
                cve_idx = headers.index("CVE漏洞編號")
            except ValueError:
                return None

            history = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[ip_idx]:
                    ip = str(row[ip_idx]).strip()
                    history[ip] = {
                        "port": str(row[port_idx]) if row[port_idx] else "",
                        "cve": str(row[cve_idx]) if row[cve_idx] else ""
                    }
            wb.close()
            return history
        except Exception as e:
            print(f"讀取歷史失敗: {e}")
            return None

    def style_header(self, ws):
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    def auto_adjust_width(self, ws):
        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_len + 2, 60)

    def generate_report(self, results):
        history = self.load_history_data()
        wb = openpyxl.Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        # 1. 掃描摘要
        ws_summary = wb.create_sheet("掃描摘要", 0)
        ws_summary.append(["項目", "內容"])
        
        summary_data = [
            ["掃描日期", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["掃描 IP 總數", len(results)],
            ["發現 CVE 總數", sum(len(r["CVE漏洞"]) for r in results)],
            ["說明", "本報告由 EASM Light Scanner 自動產生"]
        ]
        for row in summary_data: ws_summary.append(row)
        self.style_header(ws_summary)
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 40

        # 2. 詳細結果 (維持不變)
        ws_detail = wb.create_sheet("詳細結果")
        ws_detail.append(["IP", "資產狀態", "異動摘要", "地理位置", "ISP", "開放Port", "HSTS狀態", "SSL/TLS風險", "CVE漏洞編號", "建議與總結"])

        diff_rows = []
        all_ports = []

        for res in results:
            ip = res.get("IP")
            changes = []
            
            # 比對邏輯
            if not history:
                res["資產狀態"] = "🆕 首次掃描"
            elif ip not in history:
                res["資產狀態"] = "🆕 新增資產"
                changes.append("新發現主機")
            else:
                old_data = history[ip]
                current_port_set = set(res["開放Port"])
                old_port_list = old_data["port"].split(', ') if old_data["port"] and old_data["port"] != "None" else []
                old_port_set = set(old_port_list)

                added = current_port_set - old_port_set
                removed = old_port_set - current_port_set
                
                if added: changes.append(f"Port新增: {','.join(added)}")
                if removed: changes.append(f"Port關閉: {','.join(removed)}")
                
                new_cve_count = len(res["CVE漏洞"])
                old_cve_count = len(old_data["cve"].split(', ')) if old_data["cve"] and old_data["cve"] != "無" else 0
                
                if new_cve_count > old_cve_count:
                    changes.append(f"⚠️ 漏洞增加 ({new_cve_count - old_cve_count})")
                elif new_cve_count < old_cve_count:
                    changes.append(f"✅ 漏洞減少 ({old_cve_count - new_cve_count})")

            res["異動摘要"] = "; ".join(changes) if changes else "無顯著異動"
            if changes: diff_rows.append([ip, res["資產狀態"], res["異動摘要"]])
            
            all_ports.extend(res["raw_ports"])

            ws_detail.append([
                res["IP"], res["資產狀態"], res["異動摘要"], res["地理位置"], res["ISP"],
                ", ".join(res["開放Port"]), res["HSTS狀態"],
                "\n".join(res["SSL/TLS風險"]) if res["SSL/TLS風險"] else "Pass",
                ", ".join(sorted(list(res["CVE漏洞"]))) if res["CVE漏洞"] else "無",
                "; ".join(res["建議"])
            ])
        
        self.style_header(ws_detail)
        self.auto_adjust_width(ws_detail)

        # 3. 漏洞清單 (重要修正：改為條列式)
        ws_vuln = wb.create_sheet("漏洞清單")
        # 增加「Port」與「服務」欄位
        ws_vuln.append(["IP", "Port", "服務/版本", "CVE 編號", "說明"])
        
        has_vuln = False
        for res in results:
            if res["vuln_details"]: # 如果有詳細漏洞資訊
                has_vuln = True
                # 針對每一個被記錄下來的漏洞進行迭代
                for v in res["vuln_details"]:
                    ws_vuln.append([
                        res["IP"],
                        v["port"],     # 精確的 Port
                        v["service"],  # 包含版本號的服務名稱
                        v["cve"],      # CVE 編號
                        "建議更新服務版本至最新版" # 通用建議
                    ])
        
        if not has_vuln:
            ws_vuln.append(["本次掃描", "-", "-", "無", "恭喜！未發現 CVE 漏洞"])

        self.style_header(ws_vuln)
        self.auto_adjust_width(ws_vuln)

        # 4. Port 統計
        ws_stats = wb.create_sheet("Port 統計")
        ws_stats.append(["Port", "開啟主機數", "佔比"])
        if all_ports:
            port_counts = Counter(all_ports)
            for port, count in port_counts.most_common():
                ws_stats.append([int(port), count, f"{(count / len(results) * 100):.1f}%"])
        self.style_header(ws_stats)

        # 5. 變動比對
        ws_diff = wb.create_sheet("變動比對摘要")
        ws_diff.append(["IP", "狀態", "變動內容"])
        for row in diff_rows: ws_diff.append(row)
        if not diff_rows: ws_diff.append(["本次", "平穩", "無變動"])
        self.style_header(ws_diff)
        self.auto_adjust_width(ws_diff)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(OUTPUT_DIR, f"EASM_Report_v3_{timestamp}.xlsx")
        wb.save(filename)
        print(f"\n✨ 完整報表 已產生: {filename}")


    def run(self, input_file):
        if not os.path.exists(input_file):
            print(f"❌ 錯誤：找不到輸入檔案 '{input_file}'")
            return
            
        with open(input_file, 'r') as f:
            ips = [line.strip() for line in f if line.strip() and not line.startswith('#')]
        
        print(f"🚀 啟動掃描，目標：{len(ips)} 個 IP")
        results = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_ip = {executor.submit(self.scan_single_ip, ip): ip for ip in ips}
            for future in as_completed(future_to_ip):
                results.append(future.result())
        
        # 呼叫你原本的 generate_report
        # self.generate_report(results)
        print("✅ 掃描任務結束")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="整合版 EASM 掃描器")
    parser.add_argument("ip_list", help="IP 清單檔案")
    args = parser.parse_args()
    
    scanner = LightEASMScannerV3()
    scanner.run(args.ip_list) # 修正點：顯式傳遞參數
